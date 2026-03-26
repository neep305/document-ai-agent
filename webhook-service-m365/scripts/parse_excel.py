#!/usr/bin/env python3
"""
parse_excel.py — Auto-detect header row and parse Excel sheet into JSON rows.

Input  (stdin): JSON { "base64": "<base64-encoded xlsx>", "sheetName": "<name or null>" }
Output (stdout): JSON {
    "success": true,
    "sheet": "<sheet name used>",
    "header_row": <1-based int>,
    "data_start_row": <1-based int>,
    "columns": { "<col_index>": "<header_name>" },
    "rows": [ { "<header_name>": "<cell_value>", ... }, ... ]
}
On error: JSON { "success": false, "error": "<message>", "detected_sheets": [...] }
"""

import sys
import json
import base64
import io
import os

import openpyxl


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ALIASES_PATH = os.path.join(SCRIPT_DIR, "config", "column_aliases.json")

with open(ALIASES_PATH, encoding="utf-8") as f:
    ALIASES = json.load(f)

DISCOVERY_KEYWORDS = [kw.lower() for kw in ALIASES.get("discovery_header_keywords", [])]
MAX_HEADER_SCAN_ROWS = 25
MAX_DATA_ROWS = 500


def _fill_merged_cells(ws) -> None:
    """
    Propagate the master-cell value to every sibling cell in each merged range.
    openpyxl stores only the top-left (master) cell's value; all other cells in
    the range return None.  This must be called with read_only=False.
    """
    # Snapshot first — mutating cells while iterating merged_cells is unsafe
    ranges_snapshot = [
        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rng in ws.merged_cells.ranges
    ]
    for min_row, min_col, max_row, max_col in ranges_snapshot:
        master_value = ws.cell(min_row, min_col).value
        if master_value is None:
            continue  # nothing to propagate
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue  # skip master cell itself
                ws.cell(row, col).value = master_value


def _cell_non_empty(cell) -> bool:
    return cell.value is not None and str(cell.value).strip() not in ('', 'None')


def _keyword_match(val: str, keywords: list[str]) -> bool:
    v = val.strip().lower()
    return any(kw in v or v in kw for kw in keywords)


def scan_header_row(ws, keywords: list[str], max_rows: int = MAX_HEADER_SCAN_ROWS) -> int | None:
    """
    Strategy 1 (preferred): row with the most non-empty cells in the scan window.
    Strategy 2 (fallback):  first row where any cell matches a discovery keyword.
    Ignores rows above the best-count row that are clearly just title/metadata.
    """
    row_counts: list[tuple[int, int]] = []
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        count = sum(1 for cell in ws[row_idx] if _cell_non_empty(cell))
        row_counts.append((row_idx, count))

    if not row_counts:
        return None

    best_row, best_count = max(row_counts, key=lambda x: x[1])
    if best_count >= 2:
        return best_row

    # Fallback: keyword scan
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        for cell in ws[row_idx]:
            if cell.value and _keyword_match(str(cell.value), keywords):
                return row_idx
    return None


def row_to_dict(ws, row_idx: int, col_map: dict[int, str]) -> dict:
    record = {}
    for col_idx, header in col_map.items():
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is not None:
            record[header] = str(val).strip() if isinstance(val, str) else val
        else:
            record[header] = ""
    return record


def is_empty_row(record: dict) -> bool:
    return all(v == "" or v is None for v in record.values())


def parse_sheet(ws, keywords: list[str]) -> dict:
    _fill_merged_cells(ws)  # propagate merged-cell values before any parsing
    header_row = scan_header_row(ws, keywords)
    if header_row is None:
        return {"success": False, "error": f"No header row found in sheet '{ws.title}' within first {MAX_HEADER_SCAN_ROWS} rows"}

    # Build column map from header row
    col_map: dict[int, str] = {}
    for cell in ws[header_row]:
        if cell.value:
            col_map[cell.column] = str(cell.value).strip()

    if not col_map:
        return {"success": False, "error": f"Header row {header_row} has no non-empty cells"}

    data_start = header_row + 1
    rows = []
    empty_streak = 0

    for row_idx in range(data_start, min(data_start + MAX_DATA_ROWS, ws.max_row + 1)):
        record = row_to_dict(ws, row_idx, col_map)
        if is_empty_row(record):
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0
        rows.append(record)

    return {
        "success": True,
        "sheet": ws.title,
        "header_row": header_row,
        "data_start_row": data_start,
        "columns": {str(k): v for k, v in col_map.items()},
        "rows": rows,
    }


def main():
    raw = sys.stdin.buffer.read()
    try:
        payload = json.loads(raw)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Invalid input JSON: {e}"}))
        sys.exit(1)

    b64 = payload.get("base64", "")
    sheet_name = payload.get("sheetName")
    keywords = payload.get("keywords", DISCOVERY_KEYWORDS)

    try:
        xlsx_bytes = base64.b64decode(b64)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"base64 decode failed: {e}"}))
        sys.exit(1)

    try:
        # read_only=False is required so ws.merged_cells is accessible
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Cannot open workbook: {e}"}))
        sys.exit(1)

    sheet_names = wb.sheetnames

    if sheet_name:
        # Exact match first, then case-insensitive
        ws = wb[sheet_name] if sheet_name in sheet_names else None
        if ws is None:
            match = next((s for s in sheet_names if s.lower() == sheet_name.lower()), None)
            ws = wb[match] if match else None
        if ws is None:
            print(json.dumps({"success": False, "error": f"Sheet '{sheet_name}' not found", "detected_sheets": sheet_names}))
            sys.exit(1)
        result = parse_sheet(ws, keywords)
    else:
        # Try all sheets, return first success
        result = None
        for name in sheet_names:
            ws = wb[name]
            r = parse_sheet(ws, keywords)
            if r.get("success") and r.get("rows"):
                result = r
                break
        if result is None:
            print(json.dumps({"success": False, "error": "No matching sheet found", "detected_sheets": sheet_names}))
            sys.exit(1)

    wb.close()
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
