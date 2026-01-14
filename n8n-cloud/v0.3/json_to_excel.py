#!/usr/bin/env python3
"""
BRD to SDR - JSON to Excel Converter

This script:
1. Calls the n8n webhook with Excel file
2. Receives JSON response with SDR data
3. Writes SDR data to the original Excel file
4. Saves the completed Excel file

Usage:
    python3 json_to_excel.py --input AA_BRD_SDR_Test_01122026.xlsx --client "eCommerce Client A"
"""

import argparse
import base64
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import requests
except ImportError:
    print("❌ Error: 'requests' module not found")
    print("Install: pip3 install requests")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Error: 'openpyxl' module not found")
    print("Install: pip3 install openpyxl")
    sys.exit(1)


def read_and_encode_file(file_path):
    """Read Excel file and encode to base64"""
    print(f"📂 Reading file: {file_path}")

    if not Path(file_path).exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    with open(file_path, 'rb') as f:
        file_data = f.read()

    base64_data = base64.b64encode(file_data).decode('utf-8')
    file_size = len(file_data) / 1024  # KB

    print(f"✅ File read successfully ({file_size:.1f} KB)")
    print(f"📊 Base64 length: {len(base64_data)} characters")

    return base64_data


def call_n8n_webhook(webhook_url, client_name, file_base64):
    """Call n8n webhook and get SDR JSON response"""
    print(f"\n🌐 Calling n8n webhook...")
    print(f"   URL: {webhook_url}")
    print(f"   Client: {client_name}")

    payload = {
        "clientName": client_name,
        "fileData": file_base64
    }

    try:
        response = requests.post(
            webhook_url,
            json=payload,
            headers={'Content-Type': 'application/json'},
            timeout=120  # 2 minutes timeout for AI processing
        )

        print(f"📥 Response status: {response.status_code}")

        if response.status_code != 200:
            print(f"❌ Error response: {response.text}")
            raise Exception(f"HTTP {response.status_code}: {response.text}")

        result = response.json()

        if not result.get('success'):
            raise Exception(f"Webhook returned success=false: {result}")

        print(f"✅ Webhook call successful")
        print(f"📊 Stats: {result.get('stats', {})}")

        return result

    except requests.exceptions.Timeout:
        print("❌ Request timeout (120 seconds)")
        print("   The AI processing might be taking longer than expected.")
        raise
    except requests.exceptions.RequestException as e:
        print(f"❌ Request error: {e}")
        raise


def write_sdr_to_excel(input_file, sdr_data, output_file):
    """Write SDR data to Excel file"""
    print(f"\n📝 Writing SDR data to Excel...")
    print(f"   Input: {input_file}")
    print(f"   Output: {output_file}")

    # Load workbook
    wb = load_workbook(input_file)

    # Check required sheets exist
    required_sheets = ['eVars', 'Props', 'Events']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Required sheet '{sheet_name}' not found in Excel file")

    # Write eVars
    print("   Writing eVars...")
    ws_evars = wb['eVars']
    row = 7  # Start from row 7
    for evar in sdr_data.get('evars', []):
        ws_evars.cell(row, 2).value = evar.get('Requirement ID', '')
        ws_evars.cell(row, 3).value = evar.get('Analytics Variable', '')
        ws_evars.cell(row, 4).value = evar.get('Business Name', '')
        ws_evars.cell(row, 5).value = evar.get('Business Description', '')
        ws_evars.cell(row, 6).value = evar.get('Expected Values', '')
        ws_evars.cell(row, 7).value = evar.get('Implementation Trigger', '')
        ws_evars.cell(row, 8).value = evar.get('Example Value', '')
        ws_evars.cell(row, 9).value = evar.get('Additional Notes', '')
        row += 1

    print(f"   ✅ Wrote {len(sdr_data.get('evars', []))} eVars")

    # Write Props
    print("   Writing Props...")
    ws_props = wb['Props']
    row = 7
    for prop in sdr_data.get('props', []):
        ws_props.cell(row, 2).value = prop.get('Requirement ID', '')
        ws_props.cell(row, 3).value = prop.get('Analytics Variable', '')
        ws_props.cell(row, 4).value = prop.get('Business Name', '')
        ws_props.cell(row, 5).value = prop.get('Business Description', '')
        ws_props.cell(row, 6).value = prop.get('Expected Values', '')
        ws_props.cell(row, 7).value = prop.get('Implementation Trigger', '')
        ws_props.cell(row, 8).value = prop.get('Example Value', '')
        ws_props.cell(row, 9).value = prop.get('Additional Notes', '')
        row += 1

    print(f"   ✅ Wrote {len(sdr_data.get('props', []))} Props")

    # Write Events
    print("   Writing Events...")
    ws_events = wb['Events']
    row = 7
    for event in sdr_data.get('events', []):
        ws_events.cell(row, 2).value = event.get('Requirement ID', '')
        ws_events.cell(row, 3).value = event.get('Analytics Variable', '')
        ws_events.cell(row, 4).value = event.get('Business Name', '')
        ws_events.cell(row, 5).value = event.get('Business Description', '')
        ws_events.cell(row, 6).value = event.get('Expected Values', '')
        ws_events.cell(row, 7).value = event.get('Implementation Trigger', '')
        ws_events.cell(row, 8).value = event.get('Example Value', '')
        ws_events.cell(row, 9).value = event.get('Additional Notes', '')
        row += 1

    print(f"   ✅ Wrote {len(sdr_data.get('events', []))} Events")

    # Save workbook
    wb.save(output_file)
    print(f"\n✅ Excel file saved: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Convert BRD to SDR using n8n webhook and write to Excel'
    )
    parser.add_argument(
        '--input',
        '-i',
        required=True,
        help='Input Excel file (BRD with Requirements)'
    )
    parser.add_argument(
        '--client',
        '-c',
        required=True,
        help='Client name'
    )
    parser.add_argument(
        '--output',
        '-o',
        help='Output Excel file (default: SDR_<client>_<date>.xlsx)'
    )
    parser.add_argument(
        '--webhook',
        '-w',
        default='http://54.116.8.155:5678/webhook/brd-sdr-json',
        help='n8n webhook URL (default: http://54.116.8.155:5678/webhook/brd-sdr-json)'
    )

    args = parser.parse_args()

    # Determine output file name
    if args.output:
        output_file = args.output
    else:
        timestamp = datetime.now().strftime('%Y%m%d')
        client_safe = args.client.replace(' ', '_').replace('/', '_')
        output_file = f"SDR_{client_safe}_{timestamp}.xlsx"

    print("=" * 60)
    print("🚀 BRD to SDR - JSON to Excel Converter")
    print("=" * 60)
    print(f"Input file: {args.input}")
    print(f"Client: {args.client}")
    print(f"Output file: {output_file}")
    print(f"Webhook: {args.webhook}")
    print("=" * 60)

    try:
        # Step 1: Read and encode file
        file_base64 = read_and_encode_file(args.input)

        # Step 2: Call n8n webhook
        print(f"\n⏳ Please wait 30-60 seconds for AI processing...")
        result = call_n8n_webhook(args.webhook, args.client, file_base64)

        # Step 3: Extract SDR data
        if 'sdr' not in result:
            raise ValueError("Response does not contain 'sdr' field")

        sdr_data = result['sdr']

        # Validate SDR data
        if not sdr_data.get('evars'):
            raise ValueError("SDR data missing 'evars'")
        if not sdr_data.get('props'):
            raise ValueError("SDR data missing 'props'")
        if not sdr_data.get('events'):
            raise ValueError("SDR data missing 'events'")

        # Step 4: Write to Excel
        write_sdr_to_excel(args.input, sdr_data, output_file)

        # Success summary
        print("\n" + "=" * 60)
        print("🎉 SUCCESS!")
        print("=" * 60)
        print(f"📄 Output file: {output_file}")
        print(f"📊 Statistics:")
        print(f"   - eVars: {len(sdr_data['evars'])}")
        print(f"   - Props: {len(sdr_data['props'])}")
        print(f"   - Events: {len(sdr_data['events'])}")
        print(f"   - Total: {len(sdr_data['evars']) + len(sdr_data['props']) + len(sdr_data['events'])}")
        print("=" * 60)

        return 0

    except KeyboardInterrupt:
        print("\n\n⚠️  Interrupted by user")
        return 1
    except Exception as e:
        print(f"\n\n❌ Error: {e}")
        import traceback
        print("\nStack trace:")
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
