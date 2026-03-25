#!/usr/bin/env python3
"""
generate_excel.py — Write SDR data back into original Excel template,
using header auto-detection so no sheet/row/column offsets are hardcoded.

Input  (stdin): JSON {
    "base64": "<base64-encoded original xlsx>",
    "clientName": "<string>",
    "sdrData": {
        "evars":  [ {field: value, ...}, ... ],
        "props":  [ ... ],
        "events": [ ... ]
    }
}
Output (stdout): JSON {
    "success": true,
    "base64": "<base64-encoded modified xlsx>",
    "stats": { "evars": <int>, "props": <int>, "events": <int> }
}
On error: JSON { "success": false, "error": "<message>" }
"""

import sys
import json
import base64
import io
import os

import openpyxl
from openpyxl.utils import get_column_letter


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ALIASES_PATH = os.path.join(SCRIPT_DIR, "config", "column_aliases.json")

with open(ALIASES_PATH, encoding="utf-8") as f:
    ALIASES = json.load(f)

SHEET_PATTERNS: dict[str, list[str]] = {
    k: [p.lower() for p in v]
    for k, v in ALIASES.get("sheet_patterns", {}).items()
}

MAX_HEADER_SCAN_ROWS = 20
MAX_DATA_ROWS = 2000


# ── Helpers ──────────────────────────────────────────────────────────────────

def find_sheet(wb: openpyxl.Workbook, section: str):
    """Return worksheet matching section key (evars/props/events) via pattern."""
    patterns = SHEET_PATTERNS.get(section, [section.lower()])
    sheet_names = wb.sheetnames

    # Exact match first
    for name in sheet_names:
        if name.lower() in patterns:
            return wb[name]

    # Partial/contains match
    for name in sheet_names:
        nl = name.lower()
        if any(p in nl or nl in p for p in patterns):
            return wb[name]

    return None


def scan_header_row(ws, max_rows: int = MAX_HEADER_SCAN_ROWS) -> tuple[int | None, dict[int, str]]:
    """
    Find the header row by looking for the row with the most non-empty cells.
    Returns (header_row_1based, {col_index: header_name}).
    """
    best_row = None
    best_count = 0
    best_map: dict[int, str] = {}

    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        col_map = {}
        for cell in ws[row_idx]:
            if cell.value and str(cell.value).strip():
                col_map[cell.column] = str(cell.value).strip()
        if len(col_map) > best_count:
            best_count = len(col_map)
            best_row = row_idx
            best_map = col_map

    if best_count < 2:
        return None, {}
    return best_row, best_map


def build_field_to_col(col_map: dict[int, str], section: str) -> dict[str, int]:
    """
    Map SDR field names → Excel column indices using aliases config.
    Uses exact match first, then case-insensitive partial match.
    """
    field_aliases: dict[str, list[str]] = ALIASES.get(section, {})
    col_name_to_idx = {name.lower(): idx for idx, name in col_map.items()}

    mapping: dict[str, int] = {}

    for field, aliases in field_aliases.items():
        # 1. Exact match (case-insensitive)
        matched = None
        for alias in aliases:
            if alias.lower() in col_name_to_idx:
                matched = col_name_to_idx[alias.lower()]
                break

        # 2. Partial/contains match
        if matched is None:
            for alias in aliases:
                al = alias.lower()
                for col_name_lower, idx in col_name_to_idx.items():
                    if al in col_name_lower or col_name_lower in al:
                        matched = idx
                        break
                if matched is not None:
                    break

        if matched is not None:
            mapping[field] = matched

    return mapping


def get_field_value(record: dict, field: str) -> str | None:
    """Try the canonical field name plus common variations."""
    for key in [field, field.upper(), field.lower()]:
        v = record.get(key)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return None


def clear_data_rows(ws, start_row: int, col_indices: list[int]):
    """Clear cells in specified columns from start_row to max_row (skips merged cells)."""
    from openpyxl.cell.cell import MergedCell
    if start_row > ws.max_row:
        return
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in col_indices:
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = None


# ── SDR Scratch Builder ──────────────────────────────────────────────────────

# Sheet definitions: (excel_sheet_name, sdr_data_key, [(field_key, column_header), ...])
_SHEET_DEFS = [
    ("eVars", "evars", [
        ("req_id",              "Requirement ID"),
        ("variable",            "Analytics Variable"),
        ("variable_name",       "Variable Name"),
        ("variable_description","Variable Description"),
        ("value_format",        "Value Format"),
        ("allocation",          "eVar Allocation"),
        ("expiration",          "eVar Expiration"),
        ("merchandising",       "Merchandising"),
        ("capture_method",      "Capture Method"),
        ("implementation_note", "Implementation Note"),
        ("group",               "Group"),
    ]),
    ("Props", "props", [
        ("req_id",              "Requirement ID"),
        ("variable",            "Analytics Variable"),
        ("variable_name",       "Variable Name"),
        ("variable_description","Variable Description"),
        ("value_format",        "Value Format"),
        ("example_value",       "Example Value"),
        ("capture_method",      "Capture Method"),
        ("implementation_note", "Implementation Note"),
        ("group",               "Group"),
    ]),
    ("custom events (metrics)", "events", [
        ("req_id",              "Requirement ID"),
        ("event",               "Event"),
        ("event_name",          "Event Name"),
        ("event_description",   "Event Description"),
        ("event_type",          "Event Type"),
        ("related_vars",        "Related Variables"),
        ("capture_method",      "Capture Method"),
        ("implementation_note", "Implementation Note"),
        ("group",               "Group"),
    ]),
    ("Section A - OOTB", "section_a_ootb", [
        ("req_id",              "Requirement ID"),
        ("variable",            "Variable"),
        ("variable_name",       "Variable Name"),
        ("variable_description","Variable Description"),
        ("value_format",        "Value Format"),
        ("capture_method",      "Capture Method"),
        ("implementation_note", "Implementation Note"),
        ("group",               "Group"),
    ]),
]


def has_sdr_sheets(wb: openpyxl.Workbook) -> bool:
    """Return True if workbook already contains at least one SDR data sheet."""
    return any(find_sheet(wb, s) is not None for s in ("evars", "props", "events"))


def create_sdr_workbook_from_scratch(sdr_data: dict) -> tuple:
    """Create a fresh SDR workbook with all sheets and data. Returns (wb, stats)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    stats: dict[str, int] = {}
    for sheet_name, section_key, fields in _SHEET_DEFS:
        ws = wb.create_sheet(sheet_name)
        # Header row
        for col_idx, (_, header) in enumerate(fields, start=1):
            ws.cell(row=1, column=col_idx).value = header
        # Data rows
        records = sdr_data.get(section_key) or []
        for row_offset, record in enumerate(records[:MAX_DATA_ROWS]):
            for col_idx, (field_key, _) in enumerate(fields, start=1):
                value = get_field_value(record, field_key)
                if value is not None:
                    ws.cell(row=2 + row_offset, column=col_idx).value = value
        stats[section_key] = len(records)

    return wb, stats


def write_section(wb: openpyxl.Workbook, section: str, records: list[dict]) -> int:
    """Find sheet, detect headers, write records. Returns number of rows written."""
    ws = find_sheet(wb, section)
    if ws is None:
        return 0

    header_row, col_map = scan_header_row(ws)
    if header_row is None:
        return 0

    field_to_col = build_field_to_col(col_map, section)
    if not field_to_col:
        return 0

    data_start = header_row + 1
    col_indices = list(col_map.keys())
    clear_data_rows(ws, data_start, col_indices)

    written = 0
    from openpyxl.cell.cell import MergedCell
    for i, record in enumerate(records[:MAX_DATA_ROWS]):
        row_idx = data_start + i
        for field, col_idx in field_to_col.items():
            value = get_field_value(record, field)
            if value is not None:
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
        written += 1

    return written


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    raw = sys.stdin.buffer.read()
    try:
        payload = json.loads(raw)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Invalid input JSON: {e}"}))
        sys.exit(1)

    b64 = payload.get("base64", "")
    sdr_data = payload.get("sdrData", {})

    try:
        xlsx_bytes = base64.b64decode(b64)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"base64 decode failed: {e}"}))
        sys.exit(1)

    try:
        # load with keep_vba=False to avoid issues, data_only=False to preserve formulas
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Cannot open workbook: {e}"}))
        sys.exit(1)

    # If the uploaded file has no SDR sheets (e.g. it's a raw BRD), build from scratch
    if not has_sdr_sheets(wb):
        wb, scratch_stats = create_sdr_workbook_from_scratch(sdr_data)
        stats = {
            "evars":          scratch_stats.get("evars", 0),
            "props":          scratch_stats.get("props", 0),
            "events":         scratch_stats.get("events", 0),
            "section_a_ootb": scratch_stats.get("section_a_ootb", 0),
        }
    else:
        stats = {}
        for section in ("evars", "props", "events", "section_a_ootb"):
            records = sdr_data.get(section) or []
            stats[section] = write_section(wb, section, records) if records else 0

    # Serialize back to bytes
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_bytes = out_buf.getvalue()
    out_b64 = base64.b64encode(out_bytes).decode("ascii")

    print(json.dumps({"success": True, "base64": out_b64, "stats": stats}, ensure_ascii=False))


if __name__ == "__main__":
    main()
